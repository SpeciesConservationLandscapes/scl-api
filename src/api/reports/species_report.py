import shutil
import tempfile
import uuid
from pathlib import Path
import datetime

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.chart import (
    Reference,
    BarChart,
    Series
)
from openpyxl.chart.axis import DateAxis, NumericAxis

from .stats import (
    CONSERVATION,
    SPECIES_FRAGMENT,
    SURVEY,
    SURVEY_FRAGMENT,
    RESTORATION,
    RESTORATION_FRAGMENT,
    TOTAL_OCCUPIED,
    TOTAL_AVAILABLE,
    TOTAL,
)

TEMPLATE_PATH = f"{settings.BASE_DIR}/templates/scl_report_mvp.xlsx"


def generate(landscape_stats, date, species_name, species_common_name, country_name, indigenous_range, countries_dict):
    report_path = _open_new_template(TEMPLATE_PATH)
    workbook = load_workbook(filename=report_path)

    dates = list(sorted(landscape_stats.keys()))

    worksheet = workbook["country summary"]
    agg_worksheet = workbook["aggregated"]
    data_worksheet = workbook["landscapes over time"]

    _fill_aggregrated_tab(agg_worksheet, dates, landscape_stats)
    _fill_data_tab(data_worksheet, dates, landscape_stats, countries_dict)
    _replace_species_and_country_name(worksheet, species_name, species_common_name, country_name, date)
    _create_chart(worksheet, agg_worksheet, dates)
    _fill_stats_table(worksheet, landscape_stats, date, indigenous_range)
    
    # save
    workbook.save(filename=report_path)
    last_date = dates[-1]
    return report_path, last_date


def _open_new_template(template):
    """ copy template to temp file """
    tmp_dir = tempfile.gettempdir()
    tmp_file_name = f"{uuid.uuid4()}.xlsx"
    tmp_file_path = Path(tmp_dir, tmp_file_name)
    shutil.copyfile(template, tmp_file_path)
    return tmp_file_path


def _fill_aggregrated_tab(agg_worksheet, dates, landscape_stats):
    row = 2
    for landscape_type in [
        TOTAL, 
        TOTAL_OCCUPIED, 
        TOTAL_AVAILABLE,
        CONSERVATION,
        SPECIES_FRAGMENT,
        SURVEY,
        SURVEY_FRAGMENT,
        RESTORATION,
        RESTORATION_FRAGMENT]:
        for date in dates:
            agg_worksheet.cell(row=row, column=1, value=date)
            agg_worksheet.cell(row=row, column=2, value=landscape_type)
            agg_worksheet.cell(row=row, column=3, value=landscape_stats[date][landscape_type]["habitat_area"])
            agg_worksheet.cell(row=row, column=4, value=landscape_stats[date][landscape_type]["protected_area"])
            agg_worksheet.cell(row=row, column=5, value=landscape_stats[date][landscape_type]["biodiversity_area"])
            row += 1

def _fill_data_tab(data_worksheet, dates, landscape_stats, countries_dict):
    row = 2
    for landscape_type in [
        CONSERVATION,
        SPECIES_FRAGMENT,
        SURVEY,
        SURVEY_FRAGMENT,
        RESTORATION,
        RESTORATION_FRAGMENT]:
        for date in dates:
            landscapes = landscape_stats[date][landscape_type]["landscapes"]
            for landscape in landscapes:

                biome_areas = landscape["biome_areas"]
                for biome_area in (biome_areas if len(biome_areas) > 0 else [None]):

                    data_worksheet.cell(row=row, column=1, value=date)

                    data_worksheet.cell(row=row, column=2, value=landscape["country_code"])
                    data_worksheet.cell(row=row, column=3, value=countries_dict[landscape["country_code"]])
                    data_worksheet.cell(row=row, column=4, value=landscape["type"])
                    data_worksheet.cell(row=row, column=5, value=landscape["landscape_id"])
                    data_worksheet.cell(row=row, column=6, value=landscape["habitat_area"])

                    if biome_area:
                        data_worksheet.cell(row=row, column=7, value=biome_area["biomeid"])
                        data_worksheet.cell(row=row, column=8, value=biome_area["biomename"])
                        area = biome_area["protected"] + biome_area["unprotected"]
                        data_worksheet.cell(row=row, column=9, value=area)
                        data_worksheet.cell(row=row, column=10, value=biome_area["protected"])
                        data_worksheet.cell(row=row, column=11, value=biome_area["kba_area"])
                        data_worksheet.cell(row=row, column=12, value=str(biome_area["pas"]))
                        data_worksheet.cell(row=row, column=13, value=str(biome_area["kbas"]))

                    row += 1

def _replace_species_and_country_name(worksheet, species_name, species_common_name, country_name, date):
    # A1-A7 & A36
    title_range = list(zip(range(1,8), [1 for _ in range(1,8)])) + [(36, 1)]
    # B37-B39
    table_range = list(zip(range(37,40), [2 for _ in range(37,40)]))
    # A53-B63
    definition_range = list(zip(range(53,64), [1 for _ in range(53,64)])) + list(zip(range(53,64), [2 for _ in range(53,64)]))

    # Find and replace <species>, <date>, etc in various cells
    date_string = str(datetime.datetime.now().date())
    for (row, col) in title_range + table_range + definition_range:
        val = worksheet.cell(row=row, column=col).value
        if val:
            val = val.replace('<specieslatin>', species_name )
            val = val.replace('<Species>', species_common_name.capitalize()).replace('<species>', species_common_name.lower())
            val = val.replace('<Country>', country_name).replace('<date>', date).replace('<today>', date_string)
            worksheet.cell(row=row, column=col, value=val)

def _create_chart(worksheet, agg_worksheet, dates):
    # create chart
    bar_chart = BarChart()
    bar_chart.title = "Habitat changes over time"
    bar_chart.style = 12
    bar_chart.type = "col"
    bar_chart.grouping = "stacked"
    bar_chart.overlap = 100
    bar_chart.y_axis = NumericAxis()
    bar_chart.y_axis.title = "Area (km²)"
    bar_chart.y_axis.crossAx = 500
    bar_chart.y_axis.number_format = '#,###'
    bar_chart.x_axis = DateAxis(crossAx=100)
    bar_chart.x_axis.number_format = 'yyyy-mm-dd'
    bar_chart.x_axis.majorTimeUnit = "years"
    bar_chart.x_axis.title = "Assessment Date"
    bar_chart.height = 12  # default is 7.5
    bar_chart.width = 25  # default is 15

    # get references
    for stack_i, landscape_type in enumerate([
        CONSERVATION,
        SPECIES_FRAGMENT,
        SURVEY,
        SURVEY_FRAGMENT,
        RESTORATION,
        RESTORATION_FRAGMENT]):
        title = landscape_type

        ref = Reference(
            agg_worksheet, min_col=3, max_col=3, min_row=len(dates) * (3 + stack_i) + 2, max_row=len(dates) * (4 + stack_i) + 1
        )
        bar_chart.append(Series(ref, title=title))

    dates_ref = Reference(agg_worksheet, min_col=1, max_col=1,  min_row=2, max_row=len(dates) + 1)
    bar_chart.set_categories(dates_ref)

    # set style and colours
    for s1 in bar_chart.series:
        s1.graphicalProperties.line.noFill = True

    bar_chart.series[0].graphicalProperties.solidFill = "2171b5"
    bar_chart.series[1].graphicalProperties.solidFill = "6baed6"
    bar_chart.series[2].graphicalProperties.solidFill = "a63603"
    bar_chart.series[3].graphicalProperties.solidFill = "e6550d"
    bar_chart.series[4].graphicalProperties.solidFill = "fd8d3c"
    bar_chart.series[5].graphicalProperties.solidFill = "fdbe85"

    worksheet.add_chart(bar_chart, "A9")


def _fill_stats_table(worksheet, landscape_stats, date, indigenous_range):
    # Fill in stats table on main summary tab
    report_stats = landscape_stats[date]
    report_total_area = report_stats[TOTAL]["habitat_area"]
    rows_keys = [
        (38, CONSERVATION),
        (39, SPECIES_FRAGMENT),
        (40, TOTAL_OCCUPIED),
        (41, SURVEY),
        (42, SURVEY_FRAGMENT),
        (43, RESTORATION),
        (44, RESTORATION_FRAGMENT),
        (45, TOTAL_AVAILABLE),
        (46, TOTAL)
    ]
    for row, key in rows_keys:
        ls_stats = report_stats[key]
        try:
            percent_protected = round((ls_stats["protected_area"] / ls_stats["habitat_area"]) * 100., 2)
        except ZeroDivisionError:
            percent_protected = 0.
        try:
            percent_kba = round((ls_stats["biodiversity_area"] / ls_stats["habitat_area"]) * 100., 2)
        except ZeroDivisionError:
            percent_kba = 0.
        try:
            percent_habitat = round((ls_stats["habitat_area"] / report_total_area) * 100., 2)
        except ZeroDivisionError:
            percent_habitat = 0.
        try:
            percent_indigenous = round((ls_stats["habitat_area"] / indigenous_range) * 100., 2)
        except ZeroDivisionError:
            percent_indigenous = 0.
        worksheet.cell(row=row, column=3, value=ls_stats["num_landscapes"])
        worksheet.cell(row=row, column=4, value=ls_stats["habitat_area"])
        worksheet.cell(row=row, column=5, value=percent_protected)
        worksheet.cell(row=row, column=6, value=percent_kba)
        worksheet.cell(row=row, column=7, value=percent_habitat)
        worksheet.cell(row=row, column=8, value=percent_indigenous)

        # final indigenous range row - use same values form total habitat
        if key == TOTAL:
            try:
                percent_protected = round((ls_stats["protected_area"] / indigenous_range) * 100., 2)
            except ZeroDivisionError:
                percent_protected = 0.
            try:
                percent_kba = round((ls_stats["biodiversity_area"] / indigenous_range) * 100., 2)
            except ZeroDivisionError:
                percent_kba = 0.
            worksheet.cell(row=row + 1, column=4, value=indigenous_range)
            worksheet.cell(row=row + 1, column=5, value=percent_protected)
            worksheet.cell(row=row + 1, column=6, value=percent_kba)
